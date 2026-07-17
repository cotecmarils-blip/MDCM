from decimal import Decimal
import io
import json

from django.db import transaction
from django.http import HttpResponse
from rest_framework import viewsets, status
from rest_framework.decorators import action, api_view, permission_classes
from rest_framework.permissions import AllowAny, IsAuthenticated
from rest_framework.response import Response
from rest_framework.parsers import JSONParser, MultiPartParser, FormParser
import unicodedata

from .access import (
    can_manage_members,
    can_read_proyecto,
    can_write_resource,
    is_global_admin,
    user_proyecto_ids,
)
from .mcdm_utils import (
    build_alternatives_export,
    build_hierarchy_export,
    evaluate_proyecto,
    get_omoe_for_export,
)
from .sensitivity_utils import build_sensitivity_payload
from .mixins import AuthScopedViewSetMixin, scoped_queryset_for_model
from .models import ProyectoMembership


def _openpyxl():
    """Import diferido: el API arranca sin openpyxl; Excel falla con mensaje claro."""
    try:
        from openpyxl import Workbook, load_workbook
        from openpyxl.styles import Alignment, Font, PatternFill
    except ImportError as exc:
        raise ImportError(
            'Falta openpyxl. En el venv: pip install openpyxl==3.1.5'
        ) from exc
    return Workbook, load_workbook, Alignment, Font, PatternFill


def _qp_int(request, name):
    raw = request.query_params.get(name)
    if raw in (None, ''):
        return None
    try:
        return int(raw)
    except (TypeError, ValueError):
        return None

from .models import (
    Proyecto,
    Requisito,
    Alternativa,
    Capacidad,
    Caracteristica,
    CaracteristicaPlantilla,
    Documento,
    Dimension,
    Atributo,
    Subatributo,
    DocumentoCriterio,
    Escenario,
    PesoEscenario,
    Omoe,
    Mision,
    GrupoAfinidad,
    MopCriterio,
    DpCriterio,
    NodoArbol,
    VopResultado,
    InformeProyectoJob,
)
from .nodo_arbol_serializers import NodoArbolSerializer, ProyectoNivelArbolSerializer
from .arbol_nivel_service import ensure_all_ramas_niveles, ensure_niveles_arbol
from .riesgo_tabla_service import ensure_tablas_riesgo
from .omoe_serializers import (
    OmoeSerializer,
    MisionSerializer,
    GrupoAfinidadSerializer,
    MopCriterioSerializer,
    DpCriterioSerializer,
    VopResultadoSerializer,
)
from .serializers import (
    ProyectoSerializer,
    ProyectoListSerializer,
    RequisitoSerializer,
    AlternativaSerializer,
    CapacidadSerializer,
    CaracteristicaSerializer,
    CaracteristicaPlantillaSerializer,
    DocumentoSerializer,
    DimensionSerializer,
    AtributoSerializer,
    SubatributoSerializer,
    DocumentoCriterioSerializer,
    EscenarioSerializer,
    PesoEscenarioSerializer,
    PesoEscenarioItemSerializer,
)
from .validators import validate_peso_value


REQUISITO_XLSX_HEADERS = [
    'parametro',
    'valor_esperado',
    'unidad',
    'observaciones',
    'orden',
]

REQUISITO_XLSX_TITLES = {
    'parametro': 'Parámetro',
    'valor_esperado': 'Valor esperado',
    'unidad': 'Unidad',
    'observaciones': 'Observaciones',
    'orden': 'Orden',
}

REQUISITO_XLSX_ALIASES = {
    'parametro': {'parametro', 'parámetro', 'titulo', 'título', 'nombre', 'requisito', 'title'},
    'valor_esperado': {'valor_esperado', 'valor esperado', 'descripcion', 'descripción', 'value', 'expected_value'},
    'unidad': {'unidad', 'unit', 'criterio_aceptacion', 'criterio de aceptacion', 'criterio de aceptación'},
    'observaciones': {'observaciones', 'observacion', 'nota', 'notes'},
    'orden': {'orden', 'order'},
}


def _normalize_csv_value(value):
    if value is None:
        return ''
    return str(value).strip()


def _apply_requisitos_template_styles(sheet):
    _, _, Alignment, Font, PatternFill = _openpyxl()
    header_fill = PatternFill('solid', fgColor='1E3A8A')
    header_font = Font(color='FFFFFF', bold=True)
    wrap_alignment = Alignment(wrap_text=True, vertical='top')

    for index, header in enumerate(REQUISITO_XLSX_HEADERS, start=1):
        cell = sheet.cell(row=1, column=index)
        cell.value = REQUISITO_XLSX_TITLES[header]
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    sheet.freeze_panes = 'A2'
    sheet.auto_filter.ref = 'A1:E1'
    sheet.row_dimensions[1].height = 24

    widths = {
        'A': 28,
        'B': 28,
        'C': 12,
        'D': 36,
        'E': 12,
    }
    for column, width in widths.items():
        sheet.column_dimensions[column].width = width

    for row in sheet.iter_rows(min_row=2, max_row=200, min_col=1, max_col=5):
        for cell in row:
            cell.alignment = wrap_alignment


def _build_requisitos_workbook(requisitos):
    Workbook, _, Alignment, Font, _ = _openpyxl()
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = 'Requisitos'

    _apply_requisitos_template_styles(sheet)

    for row_index, requisito in enumerate(requisitos, start=2):
        values = [
            requisito.titulo,
            requisito.descripcion,
            requisito.criterio_aceptacion,
            requisito.observaciones,
            requisito.orden,
        ]
        for column_index, value in enumerate(values, start=1):
            sheet.cell(row=row_index, column=column_index, value=value)

    instructions = workbook.create_sheet('Instrucciones')
    instructions['A1'] = 'Plantilla de requisitos'
    instructions['A1'].font = Font(bold=True, size=14)
    instructions['A3'] = 'Cada fila es un requisito.'
    instructions['A4'] = 'Ejemplo: Barco | Eslora | 25 | m.'
    instructions['A5'] = 'Deja Código y Aplica a fuera; usa sólo las columnas de esta plantilla.'
    instructions['A6'] = 'Parámetro = lo que se quiere medir; Valor esperado = el valor; Unidad = m, kg, kW, etc.'
    instructions['A7'] = 'El archivo puede importarse varias veces; si usas reemplazo, se limpian los requisitos anteriores del proyecto.'
    instructions.column_dimensions['A'].width = 110
    for cell in ('A3', 'A4', 'A5', 'A6', 'A7'):
        instructions[cell].alignment = Alignment(wrap_text=True, vertical='top')

    return workbook


def _build_requisitos_xlsx(requisitos):
    buffer = io.BytesIO()
    workbook = _build_requisitos_workbook(requisitos)
    workbook.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


def _normalize_header_name(value):
    if value is None:
        return ''
    text = _normalize_csv_value(value)
    # remove diacritics and casefold
    nfkd = unicodedata.normalize('NFKD', text)
    no_diacritics = ''.join(ch for ch in nfkd if not unicodedata.combining(ch))
    return no_diacritics.casefold()


def _parse_requisitos_xlsx(file_obj):
    try:
        file_obj.seek(0)
    except Exception:
        pass

    try:
        _, load_workbook, _, _, _ = _openpyxl()
        workbook = load_workbook(file_obj, data_only=True)
    except ImportError:
        raise
    except Exception as exc:
        raise ValueError('No se pudo leer el archivo XLSX.') from exc

    if 'Requisitos' in workbook.sheetnames:
        sheet = workbook['Requisitos']
    else:
        sheet = workbook[workbook.sheetnames[0]]

    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        raise ValueError('El archivo no contiene encabezados.')

    header_row = rows[0]
    header_map = {}
    for index, value in enumerate(header_row):
        header_name = _normalize_header_name(value)
        if not header_name:
            continue
        header_map[header_name] = index

    if not header_map:
        raise ValueError('El archivo no contiene encabezados válidos.')

    parsed_rows = []
    for row_number, row_values in enumerate(rows[1:], start=2):
        row = {}
        for field, aliases in REQUISITO_XLSX_ALIASES.items():
            found = False
            for alias in aliases:
                alias_key = _normalize_header_name(alias)
                if alias_key in header_map:
                    idx = header_map[alias_key]
                    row[field] = row_values[idx] if idx < len(row_values) else ''
                    found = True
                    break
            if not found:
                row[field] = ''
        parsed_rows.append((row_number, row))
    return parsed_rows


def _row_value(row, *names):
    for name in names:
        if name in row and row[name] is not None:
            return _normalize_csv_value(row[name])
    return ''


def _upsert_requisito_from_row(proyecto, row):
    codigo = _row_value(row, 'codigo', 'Código', 'code')
    categoria = _row_value(row, 'aplica_a', 'aplica a', 'categoria', 'category', 'objeto', 'system')
    titulo = _row_value(row, 'parametro', 'parámetro', 'titulo', 'nombre', 'requisito', 'title')
    descripcion = _row_value(row, 'valor_esperado', 'valor esperado', 'descripcion', 'description')
    criterio_aceptacion = _row_value(row, 'unidad', 'unit', 'criterio_aceptacion', 'criterio de aceptacion')
    observaciones = _row_value(row, 'observaciones', 'observacion', 'notes')
    orden_raw = _row_value(row, 'orden', 'order')

    if not titulo:
        raise ValueError('Cada requisito debe tener un título.')

    orden = 0
    if orden_raw:
        try:
            orden = int(float(orden_raw))
        except ValueError as exc:
            raise ValueError(f'Orden inválido para el requisito "{titulo}".') from exc

    lookup = {'proyecto': proyecto}
    if codigo:
        lookup['codigo'] = codigo
    else:
        lookup['titulo'] = titulo

    Requisito.objects.update_or_create(
        **lookup,
        defaults={
            'titulo': titulo,
            'descripcion': descripcion,
            'categoria': categoria,
            'prioridad': Requisito.PRIORIDAD_MEDIA,
            'estado': Requisito.ESTADO_PENDIENTE,
            'criterio_aceptacion': criterio_aceptacion,
            'observaciones': observaciones,
            'orden': orden,
        },
    )


@api_view(['GET'])
def saludo(request):
    return Response({"mensaje": "Hola desde Django"})


@api_view(['GET'])
@permission_classes([AllowAny])
def fetch_drawio_draft(request, token):
    from .drawio_export import get_drawio_draft

    proyecto_id = request.query_params.get('proyecto')
    pid = int(proyecto_id) if proyecto_id else None
    xml = get_drawio_draft(token, proyecto_id=pid)
    if not xml:
        return Response(
            {'detail': 'El borrador del diagrama no existe o ya expiró.'},
            status=status.HTTP_404_NOT_FOUND,
        )
    response = HttpResponse(xml, content_type='application/vnd.jgraph.mxfile+xml; charset=utf-8')
    response['Cache-Control'] = 'no-store'
    return response


class ProyectoViewSet(AuthScopedViewSetMixin, viewsets.ModelViewSet):
    queryset = Proyecto.objects.all()
    parser_classes = (JSONParser, MultiPartParser, FormParser)

    def get_queryset(self):
        return scoped_queryset_for_model(
            self.request.user, Proyecto, self.queryset, self.request
        )

    def perform_create(self, serializer):
        proyecto = serializer.save()
        ensure_all_ramas_niveles(proyecto)
        ensure_tablas_riesgo(proyecto)
        user = self.request.user
        if user.is_authenticated and not is_global_admin(user):
            ProyectoMembership.objects.get_or_create(
                proyecto=proyecto,
                usuario=user,
                defaults={'rol': ProyectoMembership.ROL_JEFE},
            )

    @action(detail=True, methods=['get'], url_path='catalogo-dimensiones')
    def catalogo_dimensiones(self, request, pk=None):
        """Lista dimensiones de proyectos accesibles para importar como plantilla."""
        proyecto = self.get_object()
        if is_global_admin(request.user):
            from .models import Proyecto as ProyectoModel
            ids = list(ProyectoModel.objects.values_list('id', flat=True))
        else:
            ids = list(user_proyecto_ids(request.user))
        from .dimension_clone_service import listar_catalogo_dimensiones

        items = listar_catalogo_dimensiones(
            proyecto_ids=ids,
            excluir_proyecto_id=proyecto.id,
            incluir_proyecto_actual=True,
        )
        return Response({'items': items, 'count': len(items)})

    @action(detail=True, methods=['post'], url_path='importar-dimension')
    def importar_dimension(self, request, pk=None):
        """Importa (clona) una dimensión y su árbol micro desde otra (o la misma) proyectada."""
        from django.core.exceptions import ValidationError as DjangoValidationError

        from .dimension_clone_service import clonar_dimension_en_proyecto
        from .models import Omoe
        from .omoe_serializers import OmoeSerializer

        proyecto = self.get_object()
        if not can_write_resource(request.user, proyecto, 'omoe'):
            return Response(
                {'detail': 'Sin permiso para importar dimensiones en este proyecto.'},
                status=status.HTTP_403_FORBIDDEN,
            )

        fuente_id = request.data.get('fuente_omoe_id') or request.data.get('omoe_id')
        try:
            fuente_id = int(fuente_id)
        except (TypeError, ValueError):
            return Response(
                {'fuente_omoe_id': ['Indique el id de la dimensión origen.']},
                status=status.HTTP_400_BAD_REQUEST,
            )

        fuente = Omoe.objects.select_related('proyecto').filter(pk=fuente_id).first()
        if fuente is None:
            return Response(
                {'detail': 'Dimensión origen no encontrada.'},
                status=status.HTTP_404_NOT_FOUND,
            )
        if not can_read_proyecto(request.user, fuente.proyecto_id):
            return Response(
                {'detail': 'Sin permiso para leer la dimensión origen.'},
                status=status.HTTP_403_FORBIDDEN,
            )

        nombre = request.data.get('nombre_modelo')
        try:
            result = clonar_dimension_en_proyecto(
                fuente, proyecto, nombre_modelo=nombre,
            )
        except DjangoValidationError as exc:
            msg = exc.messages[0] if getattr(exc, 'messages', None) else str(exc)
            return Response({'detail': msg}, status=status.HTTP_400_BAD_REQUEST)

        dest = Omoe.objects.prefetch_related(*OMOE_TREE_PREFETCH).get(pk=result['omoe_id'])
        return Response(
            {
                **result,
                'omoe': OmoeSerializer(dest).data,
            },
            status=status.HTTP_201_CREATED,
        )

    @staticmethod
    def _arbol_backup_dict(backup):
        return {
            'id': backup.id,
            'nombre': backup.nombre,
            'descripcion': backup.descripcion or '',
            'omoe_nombre': backup.omoe_nombre or '',
            'rama_evaluacion': backup.rama_evaluacion or '',
            'nodos_count': backup.nodos_count,
            'creado_por': (
                getattr(backup.creado_por, 'username', None) if backup.creado_por_id else None
            ),
            'fecha_creacion': backup.fecha_creacion.isoformat() if backup.fecha_creacion else None,
        }

    @action(detail=True, methods=['get'], url_path='export-arbol')
    def export_arbol(self, request, pk=None):
        """Exporta una dimensión (árbol) a JSON re-importable (ciclo export→import)."""
        from .dimension_clone_service import serialize_dimension
        from .models import Omoe

        proyecto = self.get_object()
        omoe_id = request.query_params.get('omoe')
        qs = Omoe.objects.filter(proyecto=proyecto)
        omoe = (
            qs.filter(pk=omoe_id).first()
            if omoe_id
            else qs.order_by('orden', 'id').first()
        )
        if omoe is None:
            return Response(
                {'detail': 'No hay dimensión para exportar.'},
                status=status.HTTP_404_NOT_FOUND,
            )
        data = serialize_dimension(omoe)
        slug = (omoe.nombre_modelo or 'arbol').replace(' ', '_')[:40]
        return self._json_download_response(data, f'arbol_{slug}.json')

    @action(detail=True, methods=['post'], url_path='importar-arbol-json')
    def importar_arbol_json(self, request, pk=None):
        """Reconstruye una dimensión desde un JSON exportado (import del ciclo)."""
        from django.core.exceptions import ValidationError as DjangoValidationError

        from .dimension_clone_service import rebuild_dimension_from_data
        from .models import Omoe
        from .omoe_serializers import OmoeSerializer

        proyecto = self.get_object()
        if not can_write_resource(request.user, proyecto, 'omoe'):
            return Response(
                {'detail': 'Sin permiso para importar árboles en este proyecto.'},
                status=status.HTTP_403_FORBIDDEN,
            )

        body = request.data if isinstance(request.data, dict) else {}
        data = body.get('data')
        if data is None and body.get('omoe') is not None:
            # Permite enviar directamente el JSON exportado (sin envolver en «data»).
            data = body
        nombre = body.get('nombre_modelo')
        try:
            result = rebuild_dimension_from_data(data, proyecto, nombre_modelo=nombre)
        except DjangoValidationError as exc:
            msg = exc.messages[0] if getattr(exc, 'messages', None) else str(exc)
            return Response({'detail': msg}, status=status.HTTP_400_BAD_REQUEST)

        dest = Omoe.objects.prefetch_related(*OMOE_TREE_PREFETCH).get(pk=result['omoe_id'])
        return Response(
            {**result, 'omoe': OmoeSerializer(dest).data},
            status=status.HTTP_201_CREATED,
        )

    @action(detail=True, methods=['get', 'post'], url_path='arbol-backups')
    def arbol_backups(self, request, pk=None):
        """GET: lista copias de seguridad del proyecto. POST: crea una nueva."""
        from django.utils import timezone

        from .dimension_clone_service import serialize_dimension
        from .models import ArbolBackup, Omoe

        proyecto = self.get_object()

        if request.method == 'GET':
            items = [
                self._arbol_backup_dict(b)
                for b in ArbolBackup.objects.filter(proyecto=proyecto)
            ]
            return Response({'items': items, 'count': len(items)})

        if not can_write_resource(request.user, proyecto, 'omoe'):
            return Response(
                {'detail': 'Sin permiso para crear copias de seguridad.'},
                status=status.HTTP_403_FORBIDDEN,
            )

        try:
            omoe_id = int(request.data.get('omoe_id'))
        except (TypeError, ValueError):
            return Response(
                {'omoe_id': ['Indique la dimensión a respaldar.']},
                status=status.HTTP_400_BAD_REQUEST,
            )
        omoe = Omoe.objects.filter(pk=omoe_id, proyecto=proyecto).first()
        if omoe is None:
            return Response(
                {'detail': 'Dimensión no encontrada en este proyecto.'},
                status=status.HTTP_404_NOT_FOUND,
            )

        data = serialize_dimension(omoe)
        nombre = (request.data.get('nombre') or '').strip()
        if not nombre:
            nombre = f'{omoe.nombre_modelo} — {timezone.localtime():%Y-%m-%d %H:%M}'
        backup = ArbolBackup.objects.create(
            proyecto=proyecto,
            nombre=nombre[:200],
            descripcion=(request.data.get('descripcion') or '')[:2000],
            omoe_nombre=omoe.nombre_modelo or '',
            rama_evaluacion=omoe.rama_evaluacion or '',
            nodos_count=len(data.get('nodos') or []),
            data=data,
            creado_por=request.user if request.user.is_authenticated else None,
        )
        return Response(self._arbol_backup_dict(backup), status=status.HTTP_201_CREATED)

    @action(
        detail=True,
        methods=['post'],
        url_path=r'arbol-backups/(?P<backup_id>[0-9]+)/restaurar',
    )
    def restaurar_arbol_backup(self, request, pk=None, backup_id=None):
        """Restaura una copia de seguridad: reconstruye la dimensión en el proyecto."""
        from django.core.exceptions import ValidationError as DjangoValidationError

        from .dimension_clone_service import rebuild_dimension_from_data
        from .models import ArbolBackup, Omoe
        from .omoe_serializers import OmoeSerializer

        proyecto = self.get_object()
        if not can_write_resource(request.user, proyecto, 'omoe'):
            return Response(
                {'detail': 'Sin permiso para restaurar copias de seguridad.'},
                status=status.HTTP_403_FORBIDDEN,
            )
        backup = ArbolBackup.objects.filter(pk=backup_id, proyecto=proyecto).first()
        if backup is None:
            return Response(
                {'detail': 'Copia de seguridad no encontrada.'},
                status=status.HTTP_404_NOT_FOUND,
            )
        nombre = request.data.get('nombre_modelo') if isinstance(request.data, dict) else None
        try:
            result = rebuild_dimension_from_data(backup.data, proyecto, nombre_modelo=nombre)
        except DjangoValidationError as exc:
            msg = exc.messages[0] if getattr(exc, 'messages', None) else str(exc)
            return Response({'detail': msg}, status=status.HTTP_400_BAD_REQUEST)

        dest = Omoe.objects.prefetch_related(*OMOE_TREE_PREFETCH).get(pk=result['omoe_id'])
        return Response(
            {**result, 'omoe': OmoeSerializer(dest).data},
            status=status.HTTP_201_CREATED,
        )

    @action(
        detail=True,
        methods=['delete'],
        url_path=r'arbol-backups/(?P<backup_id>[0-9]+)',
    )
    def eliminar_arbol_backup(self, request, pk=None, backup_id=None):
        """Elimina una copia de seguridad."""
        from .models import ArbolBackup

        proyecto = self.get_object()
        if not can_write_resource(request.user, proyecto, 'omoe'):
            return Response(
                {'detail': 'Sin permiso para eliminar copias de seguridad.'},
                status=status.HTTP_403_FORBIDDEN,
            )
        deleted, _ = ArbolBackup.objects.filter(pk=backup_id, proyecto=proyecto).delete()
        if not deleted:
            return Response(
                {'detail': 'Copia de seguridad no encontrada.'},
                status=status.HTTP_404_NOT_FOUND,
            )
        return Response(status=status.HTTP_204_NO_CONTENT)

    @action(detail=True, methods=['get', 'put'], url_path='niveles-arbol')
    def niveles_arbol(self, request, pk=None):
        from django.core.exceptions import ValidationError as DjangoValidationError

        from .arbol_nivel_service import (
            list_niveles_arbol,
            list_niveles_arbol_por_ramas,
            save_niveles_arbol,
        )
        from .evaluacion_rama_choices import RAMAS_DIMENSION
        from .tipo_dimension_service import assert_codigo_activo, codigos_tipos_activos

        proyecto = self.get_object()
        rama_query = (request.query_params.get('rama') or '').strip().lower()
        codigos_ok = set(codigos_tipos_activos()) | set(RAMAS_DIMENSION)

        if request.method == 'GET':
            if rama_query:
                if rama_query not in codigos_ok and not proyecto.niveles_arbol.filter(
                    rama_evaluacion=rama_query,
                ).exists():
                    return Response(
                        {'rama': ['Tipo de dimensión no reconocido.']},
                        status=status.HTTP_400_BAD_REQUEST,
                    )
                ensure_niveles_arbol(proyecto, rama_query)
                niveles_qs = list_niveles_arbol(proyecto, rama_query)
                return Response(ProyectoNivelArbolSerializer(niveles_qs, many=True).data)

            por_ramas = list_niveles_arbol_por_ramas(proyecto)
            return Response({
                rama: ProyectoNivelArbolSerializer(niveles, many=True).data
                for rama, niveles in por_ramas.items()
            })

        if not can_manage_members(request.user, proyecto.id):
            return Response(
                {'detail': 'Sin permiso para configurar los niveles del árbol.'},
                status=status.HTTP_403_FORBIDDEN,
            )

        rama_body = (request.data.get('rama') or rama_query or '').strip().lower()
        try:
            assert_codigo_activo(rama_body)
        except Exception:
            if rama_body not in codigos_ok and not proyecto.niveles_arbol.filter(
                rama_evaluacion=rama_body,
            ).exists():
                return Response(
                    {'rama': ['Indique un tipo de dimensión válido del catálogo.']},
                    status=status.HTTP_400_BAD_REQUEST,
                )
        items = request.data.get('niveles')
        if not isinstance(items, list):
            return Response(
                {'niveles': ['Se esperaba una lista de niveles.']},
                status=status.HTTP_400_BAD_REQUEST,
            )
        try:
            niveles = save_niveles_arbol(proyecto, rama_body, items)
        except DjangoValidationError as exc:
            msg = exc.messages[0] if getattr(exc, 'messages', None) else str(exc)
            return Response({'detail': msg}, status=status.HTTP_400_BAD_REQUEST)
        return Response(ProyectoNivelArbolSerializer(niveles, many=True).data)

    @action(detail=True, methods=['get', 'put'], url_path='tablas-riesgo')
    def tablas_riesgo(self, request, pk=None):
        from .riesgo_tabla_service import get_tablas_riesgo_payload, save_tablas_riesgo

        proyecto = self.get_object()
        if request.method == 'GET':
            return Response(get_tablas_riesgo_payload(proyecto))
        if not can_manage_members(request.user, proyecto.id):
            return Response(
                {'detail': 'Sin permiso para configurar las tablas de riesgo.'},
                status=status.HTTP_403_FORBIDDEN,
            )
        return Response(save_tablas_riesgo(proyecto, request.data))

    def get_serializer_class(self):
        if self.action == 'list':
            return ProyectoListSerializer
        return ProyectoSerializer

    def perform_destroy(self, instance):
        """Elimina el proyecto y, por CASCADE en BD, todo lo asociado al mismo."""
        with transaction.atomic():
            instance.delete()

    def destroy(self, request, *args, **kwargs):
        body = request.data if isinstance(request.data, dict) else {}
        password = (body.get('password') or '').strip()
        if not password:
            return Response(
                {'password': ['Indique su contraseña para confirmar la eliminación.']},
                status=status.HTTP_400_BAD_REQUEST,
            )
        if not request.user.check_password(password):
            return Response(
                {'password': ['Contraseña incorrecta.']},
                status=status.HTTP_403_FORBIDDEN,
            )
        return super().destroy(request, *args, **kwargs)

    @action(detail=True, methods=['get'], url_path='requisitos/export')
    def export_requisitos(self, request, pk=None):
        proyecto = self.get_object()
        xlsx_data = _build_requisitos_xlsx(proyecto.requisitos.all())
        response = HttpResponse(
            xlsx_data,
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = f'attachment; filename="proyecto_{proyecto.pk}_requisitos.xlsx"'
        return response

    @action(detail=True, methods=['get'], url_path='requisitos/template')
    def requisitos_template(self, request, pk=None):
        xlsx_data = _build_requisitos_xlsx([])
        response = HttpResponse(
            xlsx_data,
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = 'attachment; filename="plantilla_requisitos.xlsx"'
        return response

    @action(detail=True, methods=['post'], url_path='requisitos/import')
    def import_requisitos(self, request, pk=None):
        proyecto = self.get_object()
        archivo = request.FILES.get('archivo')
        if not archivo:
            return Response(
                {'archivo': ['Debes adjuntar un archivo XLSX.']},
                status=status.HTTP_400_BAD_REQUEST,
            )

        replace = str(request.data.get('replace', 'true')).lower() not in {'0', 'false', 'no'}

        try:
            rows = _parse_requisitos_xlsx(archivo)
        except ValueError as exc:
            return Response({'detail': str(exc)}, status=status.HTTP_400_BAD_REQUEST)

        errors = []
        created_or_updated = 0
        with transaction.atomic():
            if replace:
                proyecto.requisitos.all().delete()
            for line_number, row in rows:
                try:
                    _upsert_requisito_from_row(proyecto, row)
                    created_or_updated += 1
                except ValueError as exc:
                    errors.append(f'Línea {line_number}: {exc}')

            if errors:
                transaction.set_rollback(True)
                return Response(
                    {'detail': '\n'.join(errors)},
                    status=status.HTTP_400_BAD_REQUEST,
                )

        return Response({
            'detail': 'Requisitos importados correctamente.',
            'count': created_or_updated,
            'replace': replace,
        })

    def _json_download_response(self, data, filename):
        payload = json.dumps(data, indent=4, ensure_ascii=False)
        response = HttpResponse(payload, content_type='application/json; charset=utf-8')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response

    @action(detail=True, methods=['get'], url_path='export/hierarchy')
    def export_hierarchy(self, request, pk=None):
        proyecto = self.get_object()
        omoe_id = request.query_params.get('omoe')
        omoe = get_omoe_for_export(proyecto, int(omoe_id) if omoe_id else None)
        if omoe is None:
            return Response(
                {'detail': 'No hay modelo OMOE para exportar.'},
                status=status.HTTP_404_NOT_FOUND,
            )
        data = build_hierarchy_export(omoe)
        slug = (proyecto.nombre or 'proyecto').replace(' ', '_')[:40]
        return self._json_download_response(data, f'{slug}_hierarchy.json')

    @action(detail=True, methods=['get'], url_path='export/diagram')
    def export_diagram(self, request, pk=None):
        from .drawio_export import export_proyecto_drawio

        proyecto = self.get_object()
        omoe_id = request.query_params.get('omoe')
        omoe = None
        if omoe_id:
            omoe = get_omoe_for_export(proyecto, int(omoe_id))
            if omoe is None:
                return Response(
                    {'detail': 'Dimensión OMOE no encontrada.'},
                    status=status.HTTP_404_NOT_FOUND,
                )
        xml = export_proyecto_drawio(proyecto, omoe)
        slug = (proyecto.nombre or 'proyecto').replace(' ', '_')[:40]
        response = HttpResponse(xml, content_type='application/vnd.jgraph.mxfile+xml; charset=utf-8')
        response['Content-Disposition'] = f'attachment; filename="{slug}_mapa.drawio"'
        return response

    @action(detail=True, methods=['post'], url_path='export/diagram-draft')
    def export_diagram_draft(self, request, pk=None):
        from .drawio_export import export_proyecto_drawio, store_drawio_draft

        proyecto = self.get_object()
        omoe_id = request.query_params.get('omoe')
        omoe = None
        if omoe_id:
            omoe = get_omoe_for_export(proyecto, int(omoe_id))
            if omoe is None:
                return Response(
                    {'detail': 'Dimensión OMOE no encontrada.'},
                    status=status.HTTP_404_NOT_FOUND,
                )
        xml = export_proyecto_drawio(proyecto, omoe)
        token = store_drawio_draft(
            xml,
            proyecto_id=proyecto.id,
            user_id=request.user.id,
        )
        return Response({'token': token})

    @action(detail=True, methods=['get'], url_path='export/diagram-escenarios')
    def export_diagram_escenarios(self, request, pk=None):
        from .drawio_export import export_proyecto_drawio_escenarios

        proyecto = self.get_object()
        omoe_id = request.query_params.get('omoe')
        omoe = None
        if omoe_id:
            omoe = get_omoe_for_export(proyecto, int(omoe_id))
            if omoe is None:
                return Response(
                    {'detail': 'Dimensión OMOE no encontrada.'},
                    status=status.HTTP_404_NOT_FOUND,
                )
        try:
            xml = export_proyecto_drawio_escenarios(proyecto, omoe)
        except ValueError as exc:
            return Response({'detail': str(exc)}, status=status.HTTP_400_BAD_REQUEST)
        slug = (proyecto.nombre or 'proyecto').replace(' ', '_')[:40]
        suffix = '_escenarios' if omoe is None else f'_{(omoe.nombre_modelo or "dim").replace(" ", "_")[:20]}_escenarios'
        response = HttpResponse(xml, content_type='application/vnd.jgraph.mxfile+xml; charset=utf-8')
        response['Content-Disposition'] = f'attachment; filename="{slug}{suffix}.drawio"'
        return response

    @action(detail=True, methods=['post'], url_path='export/diagram-escenarios-draft')
    def export_diagram_escenarios_draft(self, request, pk=None):
        from .drawio_export import export_proyecto_drawio_escenarios, store_drawio_draft

        proyecto = self.get_object()
        omoe_id = request.query_params.get('omoe')
        omoe = None
        if omoe_id:
            omoe = get_omoe_for_export(proyecto, int(omoe_id))
            if omoe is None:
                return Response(
                    {'detail': 'Dimensión OMOE no encontrada.'},
                    status=status.HTTP_404_NOT_FOUND,
                )
        try:
            xml = export_proyecto_drawio_escenarios(proyecto, omoe)
        except ValueError as exc:
            return Response({'detail': str(exc)}, status=status.HTTP_400_BAD_REQUEST)
        token = store_drawio_draft(
            xml,
            proyecto_id=proyecto.id,
            user_id=request.user.id,
        )
        return Response({'token': token})

    @action(detail=True, methods=['get'], url_path='export/alternatives')
    def export_alternatives(self, request, pk=None):
        proyecto = self.get_object()
        omoe_id = request.query_params.get('omoe')
        omoe = get_omoe_for_export(proyecto, int(omoe_id) if omoe_id else None)
        if omoe is None:
            return Response(
                {'detail': 'No hay modelo OMOE para exportar.'},
                status=status.HTTP_404_NOT_FOUND,
            )
        data = build_alternatives_export(proyecto, omoe)
        slug = (proyecto.nombre or 'proyecto').replace(' ', '_')[:40]
        return self._json_download_response(data, f'{slug}_alternatives.json')

    @action(detail=True, methods=['get'], url_path='analisis')
    def analisis(self, request, pk=None):
        proyecto = self.get_object()
        omoe_id = request.query_params.get('omoe')
        omoe = get_omoe_for_export(proyecto, int(omoe_id) if omoe_id else None)
        return Response(evaluate_proyecto(proyecto, omoe))

    @action(detail=True, methods=['get'], url_path='sensibilidad')
    def sensibilidad(self, request, pk=None):
        proyecto = self.get_object()
        omoe_id = request.query_params.get('omoe')
        omoe = get_omoe_for_export(proyecto, int(omoe_id) if omoe_id else None)
        selected = request.query_params.get('node', 'root')
        return Response(build_sensitivity_payload(proyecto, omoe, selected))

    @action(detail=True, methods=['get'], url_path='evaluacion/schema')
    def evaluacion_schema(self, request, pk=None):
        from .evaluacion_service import build_evaluacion_schema

        proyecto = self.get_object()
        return Response(build_evaluacion_schema(proyecto))

    @action(detail=True, methods=['get'], url_path='curvas-utilidad')
    def curvas_utilidad(self, request, pk=None):
        from .curvas_utilidad_service import build_curvas_utilidad_export

        proyecto = self.get_object()
        return Response(build_curvas_utilidad_export(proyecto))

    @action(detail=True, methods=['get'], url_path='informe-costos-word')
    def informe_costos_word(self, request, pk=None):
        """Word de costos (OMOC) con escenario Estandar."""
        from django.http import HttpResponse

        from .informe_word_service import build_informe_costos_docx

        proyecto = self.get_object()
        content = build_informe_costos_docx(proyecto)
        filename = f'informe-costos-{proyecto.id}.docx'
        response = HttpResponse(
            content,
            content_type=(
                'application/vnd.openxmlformats-officedocument.'
                'wordprocessingml.document'
            ),
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response

    @action(detail=True, methods=['get'], url_path='informe-alternativas-word')
    def informe_alternativas_word(self, request, pk=None):
        """Word de las alternativas del proyecto (fichas con fotos y anexos)."""
        from django.http import HttpResponse

        from .informe_alternativas_word_service import (
            build_informe_alternativas_docx,
        )

        proyecto = self.get_object()
        content = build_informe_alternativas_docx(proyecto)
        filename = f'alternativas-proyecto-{proyecto.id}.docx'
        response = HttpResponse(
            content,
            content_type=(
                'application/vnd.openxmlformats-officedocument.'
                'wordprocessingml.document'
            ),
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response

    @action(detail=True, methods=['get'], url_path='informe-proyecto-word')
    def informe_proyecto_word(self, request, pk=None):
        """Word integral de proyecto: alternativas, árboles/pesos y evaluaciones."""
        from django.http import HttpResponse

        from .informe_word_service import build_informe_proyecto_docx

        proyecto = self.get_object()
        include_map_weights = (
            str(request.query_params.get('map_weights', '')).strip().lower()
            in {'1', 'true', 'yes', 'si', 'sí'}
        )
        content = build_informe_proyecto_docx(
            proyecto,
            include_map_weights=include_map_weights,
        )
        filename = f'informe-proyecto-{proyecto.id}.docx'
        response = HttpResponse(
            content,
            content_type=(
                'application/vnd.openxmlformats-officedocument.'
                'wordprocessingml.document'
            ),
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response

    @action(
        detail=True,
        methods=['post'],
        url_path='informe-proyecto-word/jobs',
    )
    def iniciar_informe_proyecto_word(self, request, pk=None):
        """Inicia la generación sin bloquear la petición HTTP."""
        from .informe_job_service import start_informe_proyecto_job

        proyecto = self.get_object()
        include_map_weights = (
            str(request.data.get('map_weights', '')).strip().lower()
            in {'1', 'true', 'yes', 'si', 'sí'}
        )
        job = InformeProyectoJob.objects.create(
            proyecto=proyecto,
            usuario=request.user,
            incluir_pesos_mapas=include_map_weights,
        )
        start_informe_proyecto_job(job)
        return Response(
            {
                'job_id': str(job.id),
                'estado': job.estado,
                'progreso': job.progreso,
                'etapa': job.etapa,
            },
            status=status.HTTP_202_ACCEPTED,
        )

    @action(
        detail=True,
        methods=['get'],
        url_path='informe-proyecto-word/activo',
    )
    def informe_proyecto_word_activo(self, request, pk=None):
        """Job pendiente/en curso del usuario (para restaurar la barra al volver)."""
        proyecto = self.get_object()
        queryset = InformeProyectoJob.objects.filter(
            proyecto=proyecto,
            estado__in=[
                InformeProyectoJob.ESTADO_PENDIENTE,
                InformeProyectoJob.ESTADO_PROCESANDO,
            ],
        )
        if not is_global_admin(request.user):
            queryset = queryset.filter(usuario=request.user)
        job = queryset.order_by('-fecha_creacion').first()
        if job is None:
            return Response({'activo': False})
        return Response({
            'activo': True,
            'job_id': str(job.id),
            'estado': job.estado,
            'progreso': job.progreso,
            'etapa': job.etapa,
            'error': '',
            'listo': False,
        })

    def _informe_proyecto_job_for_user(self, request, proyecto, job_id):
        queryset = InformeProyectoJob.objects.filter(
            pk=job_id,
            proyecto=proyecto,
        )
        if not is_global_admin(request.user):
            queryset = queryset.filter(usuario=request.user)
        return queryset.first()

    @action(
        detail=True,
        methods=['get'],
        url_path=r'informe-proyecto-word/jobs/(?P<job_id>[^/.]+)',
    )
    def progreso_informe_proyecto_word(self, request, pk=None, job_id=None):
        """Devuelve porcentaje y etapa actual de la generación."""
        proyecto = self.get_object()
        job = self._informe_proyecto_job_for_user(request, proyecto, job_id)
        if job is None:
            return Response(
                {'detail': 'Generación no encontrada.'},
                status=status.HTTP_404_NOT_FOUND,
            )
        return Response({
            'job_id': str(job.id),
            'estado': job.estado,
            'progreso': job.progreso,
            'etapa': job.etapa,
            'error': job.error if job.estado == InformeProyectoJob.ESTADO_ERROR else '',
            'listo': (
                job.estado == InformeProyectoJob.ESTADO_COMPLETADO
                and bool(job.archivo)
            ),
        })

    @action(
        detail=True,
        methods=['get'],
        url_path=r'informe-proyecto-word/jobs/(?P<job_id>[^/.]+)/download',
    )
    def descargar_informe_proyecto_word(self, request, pk=None, job_id=None):
        """Descarga el archivo únicamente cuando el trabajo está completo."""
        from django.http import FileResponse

        proyecto = self.get_object()
        job = self._informe_proyecto_job_for_user(request, proyecto, job_id)
        if job is None:
            return Response(
                {'detail': 'Generación no encontrada.'},
                status=status.HTTP_404_NOT_FOUND,
            )
        if job.estado != InformeProyectoJob.ESTADO_COMPLETADO or not job.archivo:
            return Response(
                {'detail': 'El informe todavía no está listo.'},
                status=status.HTTP_409_CONFLICT,
            )
        response = FileResponse(
            job.archivo.open('rb'),
            content_type=(
                'application/vnd.openxmlformats-officedocument.'
                'wordprocessingml.document'
            ),
        )
        response['Content-Disposition'] = (
            f'attachment; filename="informe-proyecto-{proyecto.id}.docx"'
        )
        return response

    @action(detail=True, methods=['get'], url_path='informe-curvas-word')
    def informe_curvas_word(self, request, pk=None):
        """Word con curvas de utilidad finales (nodo terminal × escenario)."""
        from django.http import HttpResponse

        from .informe_word_service import build_informe_curvas_docx

        proyecto = self.get_object()
        content = build_informe_curvas_docx(proyecto)
        filename = f'informe-curvas-utilidad-{proyecto.id}.docx'
        response = HttpResponse(
            content,
            content_type=(
                'application/vnd.openxmlformats-officedocument.'
                'wordprocessingml.document'
            ),
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response

    @action(detail=True, methods=['get', 'put'], url_path='evaluacion/valores')
    def evaluacion_valores(self, request, pk=None):
        from .evaluacion_service import load_valores_map, save_valores_bulk
        from .models import Alternativa

        proyecto = self.get_object()
        alternativa_id = request.query_params.get('alternativa')
        if not alternativa_id:
            return Response(
                {'detail': 'Parámetro alternativa requerido.'},
                status=status.HTTP_400_BAD_REQUEST,
            )
        try:
            alt = Alternativa.objects.get(pk=int(alternativa_id), proyecto=proyecto)
        except (Alternativa.DoesNotExist, ValueError, TypeError):
            return Response(status=status.HTTP_404_NOT_FOUND)

        if request.method == 'GET':
            return Response({'valores': load_valores_map(alt.id)})

        if not can_write_resource(request.user, alt, 'alternativa'):
            return Response(status=status.HTTP_403_FORBIDDEN)

        raw = request.data.get('valores', request.data)
        if not isinstance(raw, dict):
            return Response(
                {'detail': 'Se esperaba un objeto valores { clave: valor }.'},
                status=status.HTTP_400_BAD_REQUEST,
            )
        saved = save_valores_bulk(alt.id, raw)
        return Response({'valores': saved})

    @action(detail=True, methods=['get'], url_path='simulacion/validar')
    def simulacion_validar(self, request, pk=None):
        from .simulacion_service import validar_simulacion

        proyecto = self.get_object()
        return Response(validar_simulacion(proyecto))

    @action(detail=True, methods=['get', 'post'], url_path='config-trazabilidad')
    def config_trazabilidad(self, request, pk=None):
        from .config_trazabilidad_service import (
            MOMENTO_ORDER,
            build_config_trazabilidad,
            registrar_sesion_config,
        )

        proyecto = self.get_object()
        omoe_id = request.query_params.get('omoe')
        omoe_pk = None
        if omoe_id:
            try:
                omoe_pk = int(omoe_id)
            except (TypeError, ValueError):
                return Response(
                    {'detail': 'Parámetro omoe inválido.'},
                    status=status.HTTP_400_BAD_REQUEST,
                )

        if request.method == 'GET':
            return Response(build_config_trazabilidad(proyecto, omoe_id=omoe_pk))

        if not can_write_resource(request.user, proyecto, 'proyecto'):
            return Response(status=status.HTTP_403_FORBIDDEN)

        body = request.data if isinstance(request.data, dict) else {}
        momento = (body.get('momento') or '').strip()
        if momento not in MOMENTO_ORDER:
            return Response(
                {'detail': 'Indique un momento válido: estructura, utilidad, pesos o evaluacion.'},
                status=status.HTTP_400_BAD_REQUEST,
            )
        notas = body.get('notas') or ''
        omoe_body = body.get('omoe_id') or body.get('omoe')
        omoe_body_pk = None
        if omoe_body not in (None, ''):
            try:
                omoe_body_pk = int(omoe_body)
            except (TypeError, ValueError):
                return Response(
                    {'detail': 'omoe_id inválido.'},
                    status=status.HTTP_400_BAD_REQUEST,
                )
        try:
            registro = registrar_sesion_config(
                proyecto,
                request.user,
                momento,
                notas=notas,
                omoe_id=omoe_body_pk,
            )
        except ValueError as exc:
            return Response({'detail': str(exc)}, status=status.HTTP_400_BAD_REQUEST)
        except Exception as exc:
            from .models import Omoe
            if isinstance(exc, Omoe.DoesNotExist):
                return Response({'detail': 'Dimensión no encontrada.'}, status=status.HTTP_404_NOT_FOUND)
            raise

        payload = build_config_trazabilidad(proyecto, omoe_id=omoe_pk)
        payload['registro'] = {
            'id': registro.id,
            'momento': registro.momento,
            'notas': registro.notas,
            'fecha': registro.fecha_creacion.isoformat(),
        }
        return Response(payload, status=status.HTTP_201_CREATED)

    @action(detail=True, methods=['get', 'post'], url_path='eventos-decision')
    def eventos_decision(self, request, pk=None):
        from .evento_decision_service import (
            consultar_auditoria,
            crear_evento,
            get_evento_activo,
            listar_colaboradores_evento,
            listar_eventos,
        )
        from .models import Omoe

        proyecto = self.get_object()

        if request.method == 'GET':
            sub = request.query_params.get('scope')
            if sub == 'activo':
                evento = get_evento_activo(proyecto.id)
                serialized = None
                if evento:
                    serialized = next(
                        (e for e in listar_eventos(proyecto) if e['id'] == evento.id),
                        None,
                    )
                return Response({'evento_activo': serialized})
            if sub == 'auditoria':
                limit_raw = _qp_int(request, 'limit') or 100
                return Response(consultar_auditoria(
                    proyecto,
                    evento_id=_qp_int(request, 'evento'),
                    omoe_id=_qp_int(request, 'omoe'),
                    participante=request.query_params.get('participante'),
                    nodo_id=_qp_int(request, 'nodo'),
                    entidad_tipo=request.query_params.get('entidad_tipo'),
                    entidad_id=_qp_int(request, 'entidad_id'),
                    tipo_cambio=request.query_params.get('tipo_cambio'),
                    limit=min(limit_raw, 500),
                ))
            if sub == 'historial-nodo':
                from .evento_decision_service import historial_entidad

                entidad_tipo = request.query_params.get('entidad_tipo') or 'nodo_arbol'
                entidad_id = _qp_int(request, 'entidad_id') or _qp_int(request, 'nodo')
                if not entidad_id:
                    return Response(
                        {'detail': 'Parámetro entidad_id o nodo requerido.'},
                        status=status.HTTP_400_BAD_REQUEST,
                    )
                return Response(historial_entidad(
                    proyecto,
                    entidad_tipo=entidad_tipo,
                    entidad_id=entidad_id,
                ))
            if sub == 'nodos-auditoria':
                from .evento_decision_service import listar_nodos_auditoria

                return Response({
                    'nodos': listar_nodos_auditoria(
                        proyecto,
                        omoe_id=_qp_int(request, 'omoe'),
                    ),
                })
            if sub == 'colaboradores':
                if not can_read_resource(request.user, proyecto, 'proyecto'):
                    return Response(status=status.HTTP_403_FORBIDDEN)
                return Response({'colaboradores': listar_colaboradores_evento(proyecto)})
            return Response({'eventos': listar_eventos(proyecto)})

        if not can_write_resource(request.user, proyecto, 'proyecto'):
            return Response(status=status.HTTP_403_FORBIDDEN)

        body = request.data if isinstance(request.data, dict) else {}
        try:
            evento = crear_evento(proyecto, request.user, body)
        except Omoe.DoesNotExist:
            return Response({'detail': 'Dimensión no encontrada.'}, status=status.HTTP_404_NOT_FOUND)
        items = listar_eventos(proyecto)
        created = next((e for e in items if e['id'] == evento.id), None)
        return Response(created, status=status.HTTP_201_CREATED)

    @action(
        detail=True,
        methods=['get', 'put', 'patch'],
        url_path=r'eventos-decision/(?P<evento_id>[0-9]+)',
    )
    def evento_decision_detail(self, request, pk=None, evento_id=None):
        from .evento_decision_service import (
            actualizar_evento,
            exportar_informe_evento,
            listar_eventos,
        )
        from .models import EventoDecision

        proyecto = self.get_object()
        try:
            evento = EventoDecision.objects.prefetch_related('participantes').get(
                pk=int(evento_id), proyecto=proyecto,
            )
        except (EventoDecision.DoesNotExist, TypeError, ValueError):
            return Response({'detail': 'Evento no encontrado.'}, status=status.HTTP_404_NOT_FOUND)

        if request.method == 'GET':
            scope = request.query_params.get('scope')
            if scope == 'informe':
                return Response(exportar_informe_evento(evento))
            return Response(next((e for e in listar_eventos(proyecto) if e['id'] == evento.id), None))

        if not can_write_resource(request.user, proyecto, 'proyecto'):
            return Response(status=status.HTTP_403_FORBIDDEN)

        body = request.data if isinstance(request.data, dict) else {}
        try:
            actualizar_evento(evento, body)
        except ValueError as exc:
            return Response({'detail': str(exc)}, status=status.HTTP_400_BAD_REQUEST)
        return Response(next((e for e in listar_eventos(proyecto) if e['id'] == evento.id), None))

    @action(
        detail=True,
        methods=['post'],
        url_path=r'eventos-decision/(?P<evento_id>[0-9]+)/activar',
    )
    def evento_decision_activar(self, request, pk=None, evento_id=None):
        from .evento_decision_service import activar_evento, listar_eventos
        from .models import EventoDecision

        proyecto = self.get_object()
        if not can_write_resource(request.user, proyecto, 'proyecto'):
            return Response(status=status.HTTP_403_FORBIDDEN)
        try:
            evento = EventoDecision.objects.get(pk=int(evento_id), proyecto=proyecto)
        except (EventoDecision.DoesNotExist, TypeError, ValueError):
            return Response({'detail': 'Evento no encontrado.'}, status=status.HTTP_404_NOT_FOUND)
        try:
            activar_evento(evento)
        except ValueError as exc:
            return Response({'detail': str(exc)}, status=status.HTTP_400_BAD_REQUEST)
        return Response(next((e for e in listar_eventos(proyecto) if e['id'] == evento.id), None))

    @action(
        detail=True,
        methods=['post'],
        url_path=r'eventos-decision/(?P<evento_id>[0-9]+)/cerrar',
    )
    def evento_decision_cerrar(self, request, pk=None, evento_id=None):
        from .evento_decision_service import cerrar_evento, listar_eventos
        from .models import EventoDecision

        proyecto = self.get_object()
        if not can_write_resource(request.user, proyecto, 'proyecto'):
            return Response(status=status.HTTP_403_FORBIDDEN)
        try:
            evento = EventoDecision.objects.get(pk=int(evento_id), proyecto=proyecto)
        except (EventoDecision.DoesNotExist, TypeError, ValueError):
            return Response({'detail': 'Evento no encontrado.'}, status=status.HTTP_404_NOT_FOUND)
        body = request.data if isinstance(request.data, dict) else {}
        try:
            cerrar_evento(evento, justificacion=body.get('justificacion_cierre') or '')
        except ValueError as exc:
            return Response({'detail': str(exc)}, status=status.HTTP_400_BAD_REQUEST)
        return Response(next((e for e in listar_eventos(proyecto) if e['id'] == evento.id), None))

    @action(detail=True, methods=['get'], url_path='simulacion/opciones')
    def simulacion_opciones(self, request, pk=None):
        from .madm_choices import simulacion_opciones_payload

        proyecto = self.get_object()
        return Response(simulacion_opciones_payload(proyecto))

    @action(detail=True, methods=['post'], url_path='simulacion/preview')
    def simulacion_preview(self, request, pk=None):
        from django.core.exceptions import ValidationError

        from .simulacion_service import preview_simulacion

        proyecto = self.get_object()
        body = request.data if isinstance(request.data, dict) else {}
        opciones = {k: v for k, v in body.items() if k != 'nombre_calculo'}
        try:
            resultado = preview_simulacion(proyecto, opciones=opciones)
        except ValidationError as exc:
            detail = exc.messages[0] if getattr(exc, 'messages', None) else str(exc)
            return Response({'detail': detail, 'ok': False, 'pasos': []})
        except Exception as exc:
            return Response({'detail': str(exc), 'ok': False, 'pasos': []})
        return Response(resultado)

    @action(detail=True, methods=['post'], url_path='simulacion/calcular')
    def simulacion_calcular(self, request, pk=None):
        from django.core.exceptions import ValidationError

        from .simulacion_service import calcular_simulacion, save_simulacion_historial

        proyecto = self.get_object()
        body = request.data if isinstance(request.data, dict) else {}
        nombre_calculo = (body.get('nombre_calculo') or '').strip()
        if not nombre_calculo:
            return Response(
                {'detail': 'Indique un nombre para el cálculo.'},
                status=status.HTTP_400_BAD_REQUEST,
            )
        opciones = {k: v for k, v in body.items() if k != 'nombre_calculo'}
        try:
            resultado = calcular_simulacion(proyecto, opciones=opciones)
        except ValidationError as exc:
            detail = exc.messages[0] if getattr(exc, 'messages', None) else str(exc)
            return Response({'detail': detail}, status=status.HTTP_400_BAD_REQUEST)
        if not resultado.get('ok'):
            return Response(resultado, status=status.HTTP_400_BAD_REQUEST)
        historial = save_simulacion_historial(
            proyecto, request.user, resultado, nombre=nombre_calculo,
        )
        resultado['historial_id'] = historial.id
        return Response(resultado)

    @action(detail=True, methods=['get'], url_path='simulacion/historial')
    def simulacion_historial_list(self, request, pk=None):
        from .simulacion_service import list_simulacion_historial

        proyecto = self.get_object()
        return Response({'items': list_simulacion_historial(proyecto)})

    @action(
        detail=True,
        methods=['get', 'delete'],
        url_path=r'simulacion/historial/(?P<historial_id>[0-9]+)',
    )
    def simulacion_historial_detail(self, request, pk=None, historial_id=None):
        from django.core.exceptions import ObjectDoesNotExist

        from .simulacion_service import (
            delete_simulacion_historial,
            get_simulacion_historial,
        )

        proyecto = self.get_object()
        try:
            hid = int(historial_id)
        except (TypeError, ValueError):
            return Response(status=status.HTTP_404_NOT_FOUND)

        if request.method == 'DELETE':
            if not can_write_resource(request.user, proyecto, 'proyecto'):
                return Response(status=status.HTTP_403_FORBIDDEN)
            if not delete_simulacion_historial(proyecto, hid):
                return Response(status=status.HTTP_404_NOT_FOUND)
            return Response(status=status.HTTP_204_NO_CONTENT)

        try:
            return Response(get_simulacion_historial(proyecto, hid))
        except ObjectDoesNotExist:
            return Response(status=status.HTTP_404_NOT_FOUND)

    @action(
        detail=True,
        methods=['get'],
        url_path=r'simulacion/historial/(?P<historial_id>[0-9]+)/informe-resultados-word',
    )
    def simulacion_informe_resultados_word(self, request, pk=None, historial_id=None):
        """Exporta el Informe de resultados (Etapa 3) del cálculo seleccionado."""
        from django.core.exceptions import ObjectDoesNotExist
        from django.http import HttpResponse

        from .informe_resultados_word_service import build_informe_resultados_from_historial

        proyecto = self.get_object()
        try:
            hid = int(historial_id)
        except (TypeError, ValueError):
            return Response(status=status.HTTP_404_NOT_FOUND)

        try:
            content = build_informe_resultados_from_historial(proyecto, hid)
        except ObjectDoesNotExist:
            return Response(
                {'detail': 'Cálculo no encontrado.'},
                status=status.HTTP_404_NOT_FOUND,
            )
        except ValueError as exc:
            return Response({'detail': str(exc)}, status=status.HTTP_400_BAD_REQUEST)

        filename = f'informe-resultados-{proyecto.id}-{hid}.docx'
        response = HttpResponse(
            content,
            content_type=(
                'application/vnd.openxmlformats-officedocument.'
                'wordprocessingml.document'
            ),
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response

    @action(detail=True, methods=['post'], url_path='simulacion/sensibilidad')
    def simulacion_sensibilidad(self, request, pk=None):
        from django.core.exceptions import ObjectDoesNotExist, ValidationError

        from .simulacion_sensibilidad_service import (
            build_sensibilidad_from_resultado,
            build_sensibilidad_model,
            build_tornado_from_resultado,
            list_sensibilidad_dimensiones,
            rank_sensibilidad_at_weights,
        )
        from .simulacion_service import get_simulacion_historial

        proyecto = self.get_object()
        body = request.data if isinstance(request.data, dict) else {}

        resultado = body.get('resultado')
        historial_id = body.get('historial_id')
        if historial_id and not resultado:
            try:
                historial = get_simulacion_historial(proyecto, int(historial_id))
                resultado = historial.get('resultado')
            except (ObjectDoesNotExist, TypeError, ValueError):
                return Response(
                    {'ok': False, 'mensaje': 'Cálculo no encontrado.'},
                    status=status.HTTP_404_NOT_FOUND,
                )

        if body.get('listar_dimensiones'):
            dims = list_sensibilidad_dimensiones(resultado or {})
            error = None if dims else 'No hay dimensiones disponibles para sensibilidad.'
            return Response({'ok': bool(dims), 'dimensiones': dims, 'mensaje': error})

        accion = (body.get('accion') or '').strip().lower()
        weights_body = body.get('weights')

        if accion == 'rank':
            if not weights_body:
                return Response(
                    {'ok': False, 'mensaje': 'Indique los pesos por dimensión.'},
                    status=status.HTTP_400_BAD_REQUEST,
                )
            try:
                payload = rank_sensibilidad_at_weights(resultado or {}, weights_body)
            except ValidationError as exc:
                detail = exc.messages[0] if getattr(exc, 'messages', None) else str(exc)
                return Response({'ok': False, 'mensaje': detail}, status=status.HTTP_400_BAD_REQUEST)
            if not payload.get('ok'):
                return Response(payload, status=status.HTTP_400_BAD_REQUEST)
            return Response(payload)

        if accion == 'tornado':
            weights_body = body.get('weights')
            try:
                payload = build_tornado_from_resultado(
                    resultado or {},
                    alternative=body.get('alternative'),
                    weights_by_dimension=weights_body if isinstance(weights_body, dict) else None,
                )
            except ValidationError as exc:
                detail = exc.messages[0] if getattr(exc, 'messages', None) else str(exc)
                return Response({'ok': False, 'mensaje': detail}, status=status.HTTP_400_BAD_REQUEST)
            if not payload.get('ok'):
                return Response(payload, status=status.HTTP_400_BAD_REQUEST)
            return Response(payload)

        if accion == 'sweep' or body.get('dimension'):
            try:
                payload = build_sensibilidad_from_resultado(
                    resultado or {},
                    dimension=body.get('dimension'),
                    pasos=body.get('pasos', 21),
                )
            except ValidationError as exc:
                detail = exc.messages[0] if getattr(exc, 'messages', None) else str(exc)
                return Response({'ok': False, 'mensaje': detail}, status=status.HTTP_400_BAD_REQUEST)
            if not payload.get('ok'):
                return Response(payload, status=status.HTTP_400_BAD_REQUEST)
            return Response(payload)

        payload = build_sensibilidad_model(resultado or {})
        if not payload.get('ok'):
            return Response(payload, status=status.HTTP_400_BAD_REQUEST)
        return Response(payload)


class RequisitoViewSet(AuthScopedViewSetMixin, viewsets.ModelViewSet):
    queryset = Requisito.objects.all()
    serializer_class = RequisitoSerializer

    def get_queryset(self):
        return scoped_queryset_for_model(
            self.request.user, Requisito, self.queryset, self.request
        )


class AlternativaViewSet(AuthScopedViewSetMixin, viewsets.ModelViewSet):
    queryset = Alternativa.objects.prefetch_related(
        'documentos', 'capacidades', 'caracteristicas__plantilla'
    ).all()
    serializer_class = AlternativaSerializer
    parser_classes = (MultiPartParser, FormParser)

    def get_queryset(self):
        return scoped_queryset_for_model(
            self.request.user, Alternativa, self.queryset, self.request
        )


class CapacidadViewSet(AuthScopedViewSetMixin, viewsets.ModelViewSet):
    queryset = Capacidad.objects.all()
    serializer_class = CapacidadSerializer

    def get_queryset(self):
        return scoped_queryset_for_model(
            self.request.user, Capacidad, self.queryset, self.request
        )


class CaracteristicaPlantillaViewSet(AuthScopedViewSetMixin, viewsets.ModelViewSet):
    queryset = CaracteristicaPlantilla.objects.all()
    serializer_class = CaracteristicaPlantillaSerializer

    def get_queryset(self):
        return scoped_queryset_for_model(
            self.request.user, CaracteristicaPlantilla, self.queryset, self.request
        )


class CaracteristicaViewSet(AuthScopedViewSetMixin, viewsets.ModelViewSet):
    queryset = Caracteristica.objects.select_related('plantilla').all()
    serializer_class = CaracteristicaSerializer

    def get_queryset(self):
        return scoped_queryset_for_model(
            self.request.user, Caracteristica, self.queryset, self.request
        )


class DocumentoViewSet(AuthScopedViewSetMixin, viewsets.ModelViewSet):
    queryset = Documento.objects.all()
    serializer_class = DocumentoSerializer
    parser_classes = (MultiPartParser, FormParser)

    def get_queryset(self):
        return scoped_queryset_for_model(
            self.request.user, Documento, self.queryset, self.request
        )


class DimensionViewSet(AuthScopedViewSetMixin, viewsets.ModelViewSet):
    queryset = Dimension.objects.prefetch_related(
        'documentos',
        'atributos__documentos',
        'atributos__subatributos__documentos',
    ).all()
    serializer_class = DimensionSerializer

    def get_queryset(self):
        return scoped_queryset_for_model(
            self.request.user, Dimension, self.queryset, self.request
        )


class AtributoViewSet(AuthScopedViewSetMixin, viewsets.ModelViewSet):
    queryset = Atributo.objects.prefetch_related('documentos', 'subatributos__documentos').all()
    serializer_class = AtributoSerializer

    def get_queryset(self):
        return scoped_queryset_for_model(
            self.request.user, Atributo, self.queryset, self.request
        )


class SubatributoViewSet(AuthScopedViewSetMixin, viewsets.ModelViewSet):
    queryset = Subatributo.objects.prefetch_related('documentos').all()
    serializer_class = SubatributoSerializer

    def get_queryset(self):
        return scoped_queryset_for_model(
            self.request.user, Subatributo, self.queryset, self.request
        )


class EscenarioViewSet(AuthScopedViewSetMixin, viewsets.ModelViewSet):
    queryset = Escenario.objects.all()
    serializer_class = EscenarioSerializer

    def get_queryset(self):
        qs = scoped_queryset_for_model(
            self.request.user, Escenario, self.queryset, self.request
        )
        omoe_id = self.request.query_params.get('omoe')
        proyecto_id = self.request.query_params.get('proyecto')
        if omoe_id:
            return qs.filter(omoe_id=omoe_id)
        if proyecto_id:
            return qs.filter(proyecto_id=proyecto_id)
        return qs

    def perform_create(self, serializer):
        from .peso_service import after_escenario_saved

        escenario = serializer.save()
        after_escenario_saved(escenario, created=True)

    def perform_update(self, serializer):
        from .peso_service import rebalance_pesos_to_100

        old_omoe_id = serializer.instance.omoe_id
        escenario = serializer.save()
        if old_omoe_id != escenario.omoe_id and old_omoe_id:
            rebalance_pesos_to_100(Escenario.objects.filter(omoe_id=old_omoe_id))

    def perform_destroy(self, instance):
        from .peso_service import rebalance_pesos_to_100

        omoe_id = instance.omoe_id
        instance.delete()
        if omoe_id:
            rebalance_pesos_to_100(Escenario.objects.filter(omoe_id=omoe_id))

    @action(detail=True, methods=['get', 'put'], url_path='pesos')
    def pesos(self, request, pk=None):
        escenario = self.get_object()
        if request.method == 'PUT' and not can_write_resource(
            request.user, escenario, 'escenario'
        ):
            return Response(status=status.HTTP_403_FORBIDDEN)
        if request.method == 'GET':
            qs = PesoEscenario.objects.filter(escenario=escenario)
            return Response(PesoEscenarioSerializer(qs, many=True).data)

        raw = request.data.get('pesos', request.data)
        if not isinstance(raw, list):
            return Response(
                {'pesos': ['Se esperaba una lista de pesos.']},
                status=status.HTTP_400_BAD_REQUEST,
            )
        item_serializer = PesoEscenarioItemSerializer(data=raw, many=True)
        item_serializer.is_valid(raise_exception=True)
        items = item_serializer.validated_data

        with transaction.atomic():
            PesoEscenario.objects.filter(escenario=escenario).delete()
            for entry in items:
                peso = validate_peso_value(entry['peso'])
                kwargs = {
                    'escenario': escenario,
                    'peso': peso,
                    'dimension': None,
                    'atributo': None,
                    'subatributo': None,
                }
                if entry['level'] == 'dimension':
                    kwargs['dimension_id'] = entry['id']
                elif entry['level'] == 'atributo':
                    kwargs['atributo_id'] = entry['id']
                else:
                    kwargs['subatributo_id'] = entry['id']
                PesoEscenario.objects.create(**kwargs)

        qs = PesoEscenario.objects.filter(escenario=escenario)
        return Response(PesoEscenarioSerializer(qs, many=True).data)

    @action(detail=True, methods=['get', 'put'], url_path='config-arbol')
    def config_arbol(self, request, pk=None):
        from django.core.exceptions import ValidationError as DjangoValidationError

        from .nodo_escenario_service import (
            get_arbol_config_payload,
            get_nodo_config_payload,
            save_arbol_config,
            save_nodo_config,
        )

        escenario = self.get_object()
        if request.method == 'PUT' and not can_write_resource(
            request.user, escenario, 'escenario'
        ):
            return Response(status=status.HTTP_403_FORBIDDEN)

        nodo_raw = request.query_params.get('nodo')
        if nodo_raw is None and isinstance(request.data, dict):
            nodo_raw = request.data.get('nodo_id')
        nodo_id = None
        if nodo_raw is not None:
            try:
                nodo_id = int(nodo_raw)
            except (TypeError, ValueError):
                return Response(
                    {'detail': 'Parámetro «nodo» inválido.'},
                    status=status.HTTP_400_BAD_REQUEST,
                )

        if nodo_id is not None:
            if request.method == 'GET':
                try:
                    payload = get_nodo_config_payload(escenario, nodo_id)
                except DjangoValidationError as exc:
                    detail = exc.messages[0] if getattr(exc, 'messages', None) else str(exc)
                    return Response({'detail': detail}, status=status.HTTP_400_BAD_REQUEST)
                return Response(payload)
            body = request.data if isinstance(request.data, dict) else {}
            try:
                payload = save_nodo_config(escenario, nodo_id, body, usuario=request.user)
            except DjangoValidationError as exc:
                msgs = list(exc.messages) if getattr(exc, 'messages', None) else [str(exc)]
                return Response({'detail': msgs[0], 'errors': msgs}, status=status.HTTP_400_BAD_REQUEST)
            return Response(payload)

        if request.method == 'GET':
            try:
                payload = get_arbol_config_payload(escenario)
            except DjangoValidationError as exc:
                detail = exc.messages[0] if getattr(exc, 'messages', None) else str(exc)
                return Response({'detail': detail}, status=status.HTTP_400_BAD_REQUEST)
            return Response(payload)

        raw = request.data.get('nodos', request.data)
        if not isinstance(raw, list):
            return Response(
                {'detail': 'Se esperaba una lista «nodos» con la configuración del árbol.'},
                status=status.HTTP_400_BAD_REQUEST,
            )
        try:
            payload = save_arbol_config(escenario, raw, usuario=request.user)
        except DjangoValidationError as exc:
            msgs = list(exc.messages) if getattr(exc, 'messages', None) else [str(exc)]
            return Response({'detail': msgs[0], 'errors': msgs}, status=status.HTTP_400_BAD_REQUEST)
        return Response(payload)

    @action(detail=True, methods=['get', 'put'], url_path='peso-grupo')
    def peso_grupo(self, request, pk=None):
        from django.core.exceptions import ValidationError as DjangoValidationError

        from .peso_grupo_ahp_service import (
            apply_grupo_pesos,
            build_grupo_payload,
            parse_parent_id,
            save_grupo_config,
        )

        escenario = self.get_object()
        if request.method == 'PUT' and not can_write_resource(
            request.user, escenario, 'escenario'
        ):
            return Response(status=status.HTTP_403_FORBIDDEN)

        try:
            parent_id = parse_parent_id(request.query_params.get('parent'))
        except DjangoValidationError as exc:
            detail = exc.messages[0] if getattr(exc, 'messages', None) else str(exc)
            return Response({'detail': detail}, status=status.HTTP_400_BAD_REQUEST)

        if request.method == 'GET':
            try:
                payload = build_grupo_payload(escenario, parent_id)
            except DjangoValidationError as exc:
                detail = exc.messages[0] if getattr(exc, 'messages', None) else str(exc)
                return Response({'detail': detail}, status=status.HTTP_400_BAD_REQUEST)
            return Response(payload)

        body = request.data if isinstance(request.data, dict) else {}
        try:
            payload = save_grupo_config(escenario, parent_id, body, usuario=request.user)
        except DjangoValidationError as exc:
            msgs = list(exc.messages) if getattr(exc, 'messages', None) else [str(exc)]
            return Response({'detail': msgs[0], 'errors': msgs}, status=status.HTTP_400_BAD_REQUEST)
        return Response(payload)

    @action(detail=True, methods=['post'], url_path='peso-grupo/aplicar')
    def peso_grupo_aplicar(self, request, pk=None):
        from django.core.exceptions import ValidationError as DjangoValidationError

        from .peso_grupo_ahp_service import apply_grupo_pesos, parse_parent_id

        escenario = self.get_object()
        if not can_write_resource(request.user, escenario, 'escenario'):
            return Response(status=status.HTTP_403_FORBIDDEN)

        try:
            parent_id = parse_parent_id(request.query_params.get('parent'))
        except DjangoValidationError as exc:
            detail = exc.messages[0] if getattr(exc, 'messages', None) else str(exc)
            return Response({'detail': detail}, status=status.HTTP_400_BAD_REQUEST)

        try:
            payload = apply_grupo_pesos(escenario, parent_id, usuario=request.user)
        except DjangoValidationError as exc:
            msgs = list(exc.messages) if getattr(exc, 'messages', None) else [str(exc)]
            return Response({'detail': msgs[0], 'errors': msgs}, status=status.HTTP_400_BAD_REQUEST)
        return Response(payload)

    @action(
        detail=True,
        methods=['get', 'put'],
        url_path=r'config-nodo/(?P<nodo_id>[0-9]+)',
    )
    def config_nodo(self, request, pk=None, nodo_id=None):
        from django.core.exceptions import ValidationError as DjangoValidationError

        from .nodo_escenario_service import get_nodo_config_payload, save_nodo_config

        escenario = self.get_object()
        if request.method == 'PUT' and not can_write_resource(
            request.user, escenario, 'escenario'
        ):
            return Response(status=status.HTTP_403_FORBIDDEN)
        try:
            nid = int(nodo_id)
        except (TypeError, ValueError):
            return Response({'detail': 'ID de nodo inválido.'}, status=status.HTTP_400_BAD_REQUEST)

        if request.method == 'GET':
            try:
                payload = get_nodo_config_payload(escenario, nid)
            except DjangoValidationError as exc:
                detail = exc.messages[0] if getattr(exc, 'messages', None) else str(exc)
                return Response({'detail': detail}, status=status.HTTP_400_BAD_REQUEST)
            return Response(payload)

        body = request.data if isinstance(request.data, dict) else {}
        try:
            payload = save_nodo_config(escenario, nid, body, usuario=request.user)
        except DjangoValidationError as exc:
            msgs = list(exc.messages) if getattr(exc, 'messages', None) else [str(exc)]
            return Response({'detail': msgs[0], 'errors': msgs}, status=status.HTTP_400_BAD_REQUEST)
        return Response(payload)


class DocumentoCriterioViewSet(AuthScopedViewSetMixin, viewsets.ModelViewSet):
    queryset = DocumentoCriterio.objects.all()
    serializer_class = DocumentoCriterioSerializer
    parser_classes = (MultiPartParser, FormParser)

    def get_queryset(self):
        qs = scoped_queryset_for_model(
            self.request.user, DocumentoCriterio, self.queryset, self.request
        )
        dimension_id = self.request.query_params.get('dimension', None)
        atributo_id = self.request.query_params.get('atributo', None)
        subatributo_id = self.request.query_params.get('subatributo', None)
        if dimension_id:
            return qs.filter(dimension_id=dimension_id)
        if atributo_id:
            return qs.filter(atributo_id=atributo_id)
        if subatributo_id:
            return qs.filter(subatributo_id=subatributo_id)
        return qs


OMOE_TREE_PREFETCH = (
    'grupos__mops__dps',
    'misiones__grupos__mops__dps',
)


class TipoDimensionViewSet(viewsets.ModelViewSet):
    """Catálogo global de tipos de dimensión (lectura autenticada; escritura admin global)."""
    queryset = None  # set in get_queryset
    serializer_class = None
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        from .models import TipoDimension
        from .tipo_dimension_service import ensure_tipos_sistema
        ensure_tipos_sistema()
        qs = TipoDimension.objects.all().order_by('orden', 'codigo', 'id')
        if self.action in ('list', 'retrieve') and self.request.query_params.get('all') != '1':
            if not is_global_admin(self.request.user):
                qs = qs.filter(activo=True)
        return qs

    def get_serializer_class(self):
        from .tipo_dimension_serializers import TipoDimensionSerializer
        return TipoDimensionSerializer

    def create(self, request, *args, **kwargs):
        if not is_global_admin(request.user):
            return Response(
                {'detail': 'Solo el administrador global puede crear tipos de dimensión.'},
                status=status.HTTP_403_FORBIDDEN,
            )
        return super().create(request, *args, **kwargs)

    def update(self, request, *args, **kwargs):
        if not is_global_admin(request.user):
            return Response(
                {'detail': 'Solo el administrador global puede editar tipos de dimensión.'},
                status=status.HTTP_403_FORBIDDEN,
            )
        return super().update(request, *args, **kwargs)

    def partial_update(self, request, *args, **kwargs):
        return self.update(request, *args, **kwargs)

    def destroy(self, request, *args, **kwargs):
        if not is_global_admin(request.user):
            return Response(
                {'detail': 'Solo el administrador global puede desactivar tipos.'},
                status=status.HTTP_403_FORBIDDEN,
            )
        instance = self.get_object()
        if instance.es_sistema:
            return Response(
                {'detail': 'Los tipos de sistema (omoe/omoc/omor) no se eliminan; desactívelos si aplica.'},
                status=status.HTTP_400_BAD_REQUEST,
            )
        # Soft-delete: marcar inactivo para no romper FKs lógicos (códigos en Omoe).
        instance.activo = False
        instance.save(update_fields=['activo', 'fecha_actualizacion'])
        return Response(status=status.HTTP_204_NO_CONTENT)


class OmoeViewSet(AuthScopedViewSetMixin, viewsets.ModelViewSet):
    queryset = Omoe.objects.prefetch_related(*OMOE_TREE_PREFETCH).all()
    serializer_class = OmoeSerializer

    def get_queryset(self):
        return scoped_queryset_for_model(
            self.request.user, Omoe, self.queryset, self.request
        )

    def perform_create(self, serializer):
        proyecto = serializer.validated_data['proyecto']
        from .arbol_nivel_service import next_orden_omoe

        omoe = serializer.save(orden=next_orden_omoe(proyecto.id))
        ensure_all_ramas_niveles(omoe.proyecto)
        from .arbol_nivel_service import ensure_niveles_arbol
        ensure_niveles_arbol(omoe.proyecto, omoe.rama_evaluacion)
        from .escenario_service import ensure_escenario_estandar

        ensure_escenario_estandar(omoe)

    def perform_update(self, serializer):
        omoe = serializer.save()
        from .models import Escenario

        rama = (omoe.rama_evaluacion or 'omoe').strip()
        if rama == 'auto':
            rama = 'omoe'
        Escenario.objects.filter(omoe=omoe).update(rama_evaluacion=rama)


class MisionViewSet(AuthScopedViewSetMixin, viewsets.ModelViewSet):
    queryset = Mision.objects.prefetch_related(
        'grupos__mops__dps',
    ).all()
    serializer_class = MisionSerializer

    def get_queryset(self):
        return scoped_queryset_for_model(
            self.request.user, Mision, self.queryset, self.request
        )

    def perform_create(self, serializer):
        parent_id = self.request.data.get('parent_id')
        if parent_id and 'omoe' not in serializer.validated_data:
            mision = serializer.save(omoe_id=parent_id)
        else:
            mision = serializer.save()
        from .peso_service import after_mision_saved

        after_mision_saved(mision, created=True)


class GrupoAfinidadViewSet(AuthScopedViewSetMixin, viewsets.ModelViewSet):
    queryset = GrupoAfinidad.objects.prefetch_related('mops__dps').all()
    serializer_class = GrupoAfinidadSerializer

    def get_queryset(self):
        return scoped_queryset_for_model(
            self.request.user, GrupoAfinidad, self.queryset, self.request
        )

    def perform_create(self, serializer):
        from .peso_service import after_grupo_afinidad_saved

        parent_id = self.request.data.get('parent_id')
        if not parent_id:
            grupo = serializer.save()
            after_grupo_afinidad_saved(grupo, created=True)
            return
        if Omoe.objects.filter(pk=parent_id).exists():
            grupo = serializer.save(omoe_id=parent_id)
        elif Mision.objects.filter(pk=parent_id).exists():
            grupo = serializer.save(mision_id=parent_id)
        else:
            grupo = serializer.save()
        after_grupo_afinidad_saved(grupo, created=True)


class MopCriterioViewSet(AuthScopedViewSetMixin, viewsets.ModelViewSet):
    queryset = MopCriterio.objects.prefetch_related('dps').all()
    serializer_class = MopCriterioSerializer

    def get_queryset(self):
        return scoped_queryset_for_model(
            self.request.user, MopCriterio, self.queryset, self.request
        )

    def perform_create(self, serializer):
        from .peso_service import after_mop_saved

        parent_id = self.request.data.get('parent_id')
        if parent_id and 'grupo_afinidad' not in serializer.validated_data:
            mop = serializer.save(grupo_afinidad_id=parent_id)
        else:
            mop = serializer.save()
        after_mop_saved(mop, created=True)


class DpCriterioViewSet(AuthScopedViewSetMixin, viewsets.ModelViewSet):
    queryset = DpCriterio.objects.all()
    serializer_class = DpCriterioSerializer

    def get_queryset(self):
        return scoped_queryset_for_model(
            self.request.user, DpCriterio, self.queryset, self.request
        )

    def perform_create(self, serializer):
        from .peso_service import after_dp_saved

        parent_id = self.request.data.get('parent_id')
        if parent_id and 'mop' not in serializer.validated_data:
            dp = serializer.save(mop_id=parent_id)
        else:
            dp = serializer.save()
        after_dp_saved(dp, created=True)


class NodoArbolViewSet(AuthScopedViewSetMixin, viewsets.ModelViewSet):
    queryset = NodoArbol.objects.select_related('tipo_nivel', 'omoe', 'parent').all()
    serializer_class = NodoArbolSerializer

    def get_queryset(self):
        return scoped_queryset_for_model(
            self.request.user, NodoArbol, self.queryset, self.request
        )

    def perform_create(self, serializer):
        from .evento_decision_service import registrar_cambio
        from .models import EventoDecisionRegistro
        from .peso_service import after_nodo_arbol_saved

        parent_id = self.request.data.get('parent_id')
        if parent_id and 'parent' not in serializer.validated_data:
            if Omoe.objects.filter(pk=parent_id).exists():
                nodo = serializer.save(omoe_id=parent_id, parent=None)
                after_nodo_arbol_saved(nodo, created=True)
            else:
                parent = (
                    NodoArbol.objects.filter(pk=parent_id)
                    .select_related('omoe')
                    .first()
                )
                if parent:
                    nodo = serializer.save(parent_id=parent_id, omoe_id=parent.omoe_id)
                    after_nodo_arbol_saved(nodo, created=True)
                else:
                    nodo = serializer.save()
                    after_nodo_arbol_saved(nodo, created=True)
        else:
            nodo = serializer.save()
            after_nodo_arbol_saved(nodo, created=True)
        if nodo.omoe_id:
            registrar_cambio(
                nodo.omoe.proyecto_id,
                self.request.user,
                tipo_cambio=EventoDecisionRegistro.TIPO_ESTRUCTURA,
                entidad_tipo='nodo_arbol',
                entidad_id=nodo.id,
                entidad_nombre=nodo.nombre,
                campo='creacion',
                valor_anterior=None,
                valor_nuevo={'nombre': nodo.nombre, 'tipo_nivel_id': nodo.tipo_nivel_id},
                omoe_id=nodo.omoe_id,
                notas='Nodo creado en el árbol',
            )

    def perform_update(self, serializer):
        from .evento_decision_service import auditar_campos_instancia, registrar_cambio
        from .models import EventoDecisionRegistro

        instance = serializer.instance
        validated = serializer.validated_data
        campos_snapshot = (
            'peso', 'justificacion_peso',
            'tipo_criterio', 'familia_funciones', 'parametros_funcion',
            'modo_evaluacion', 'valor_umbral', 'valor_meta',
            'nombre', 'descripcion', 'codigo', 'aplica', 'observaciones',
        )
        valores_anteriores = {
            campo: getattr(instance, campo, None)
            for campo in campos_snapshot
        }
        # Copiar dicts mutables (JSONField) para no comparar contra la misma ref.
        for campo in ('parametros_funcion',):
            val = valores_anteriores.get(campo)
            if isinstance(val, dict):
                valores_anteriores[campo] = dict(val)

        peso_changed = 'peso' in validated
        nodo = serializer.save()
        if nodo.omoe_id:
            auditar_campos_instancia(
                nodo.omoe.proyecto_id,
                self.request.user,
                instancia=nodo,
                validated_data=validated,
                tipo_entidad='nodo_arbol',
                entidad_nombre=nodo.nombre,
                omoe_id=nodo.omoe_id,
                valores_anteriores=valores_anteriores,
            )
            for campo in ('nombre', 'descripcion', 'codigo', 'aplica', 'observaciones'):
                if campo not in validated:
                    continue
                anterior = valores_anteriores.get(campo)
                nuevo = validated[campo]
                if anterior == nuevo:
                    continue
                registrar_cambio(
                    nodo.omoe.proyecto_id,
                    self.request.user,
                    tipo_cambio=EventoDecisionRegistro.TIPO_ESTRUCTURA,
                    entidad_tipo='nodo_arbol',
                    entidad_id=nodo.id,
                    entidad_nombre=nodo.nombre,
                    campo=campo,
                    valor_anterior=anterior,
                    valor_nuevo=nuevo,
                    omoe_id=nodo.omoe_id,
                )
        if peso_changed and nodo.aplica:
            from .peso_service import fix_peso_total_to_100

            siblings_qs = NodoArbol.objects.filter(
                omoe_id=nodo.omoe_id, parent_id=nodo.parent_id, aplica=True,
            )
            fix_peso_total_to_100(siblings_qs, exclude_pk=nodo.pk)

    def perform_destroy(self, instance):
        from .peso_service import rebalance_pesos_to_100

        omoe_id = instance.omoe_id
        parent_id = instance.parent_id
        instance.delete()
        siblings_qs = NodoArbol.objects.filter(
            omoe_id=omoe_id, parent_id=parent_id, aplica=True
        )
        if siblings_qs.exists():
            rebalance_pesos_to_100(siblings_qs)

    @action(detail=True, methods=['get', 'put'], url_path='config-escenario')
    def config_escenario(self, request, pk=None):
        from django.core.exceptions import ValidationError as DjangoValidationError

        from .models import Escenario
        from .nodo_escenario_service import get_nodo_config_payload, save_nodo_config

        nodo = self.get_object()
        esc_raw = request.query_params.get('escenario')
        if esc_raw is None and isinstance(request.data, dict):
            esc_raw = request.data.get('escenario_id')
        if esc_raw is None:
            return Response(
                {'detail': 'Indique el escenario (query «escenario» o body escenario_id).'},
                status=status.HTTP_400_BAD_REQUEST,
            )
        try:
            escenario = Escenario.objects.get(pk=int(esc_raw), omoe_id=nodo.omoe_id)
        except (Escenario.DoesNotExist, TypeError, ValueError):
            return Response(
                {'detail': 'Escenario no encontrado para la dimensión de este nodo.'},
                status=status.HTTP_404_NOT_FOUND,
            )
        if request.method == 'PUT' and not can_write_resource(
            request.user, escenario, 'escenario'
        ):
            return Response(status=status.HTTP_403_FORBIDDEN)
        if request.method == 'GET':
            try:
                payload = get_nodo_config_payload(escenario, nodo.id)
            except DjangoValidationError as exc:
                detail = exc.messages[0] if getattr(exc, 'messages', None) else str(exc)
                return Response({'detail': detail}, status=status.HTTP_400_BAD_REQUEST)
            return Response(payload)
        body = request.data if isinstance(request.data, dict) else {}
        try:
            payload = save_nodo_config(escenario, nodo.id, body, usuario=request.user)
        except DjangoValidationError as exc:
            msgs = list(exc.messages) if getattr(exc, 'messages', None) else [str(exc)]
            return Response({'detail': msgs[0], 'errors': msgs}, status=status.HTTP_400_BAD_REQUEST)
        return Response(payload)

    @action(detail=False, methods=['post'], url_path='reordenar')
    def reordenar(self, request):
        ids = request.data.get('ids')
        if not isinstance(ids, list) or len(ids) < 2:
            return Response(
                {'detail': 'Se requiere una lista ids con al menos dos nodos.'},
                status=status.HTTP_400_BAD_REQUEST,
            )
        try:
            int_ids = [int(i) for i in ids]
        except (TypeError, ValueError):
            return Response({'detail': 'ids inválidos.'}, status=status.HTTP_400_BAD_REQUEST)

        nodos = list(
            self.get_queryset()
            .filter(pk__in=int_ids)
            .select_related('tipo_nivel', 'omoe', 'omoe__proyecto')
        )
        if len(nodos) != len(int_ids):
            return Response(
                {'detail': 'Uno o más nodos no existen o no tiene permiso.'},
                status=status.HTTP_404_NOT_FOUND,
            )

        from .arbol_nivel_service import reordenar_nodos_arbol

        try:
            reordenar_nodos_arbol(nodos, int_ids)
        except ValueError as exc:
            return Response({'detail': str(exc)}, status=status.HTTP_400_BAD_REQUEST)

        return Response({'ok': True, 'ids': int_ids})


class VopResultadoViewSet(AuthScopedViewSetMixin, viewsets.ModelViewSet):
    queryset = VopResultado.objects.select_related('dp', 'alternativa').all()
    serializer_class = VopResultadoSerializer

    def get_queryset(self):
        return scoped_queryset_for_model(
            self.request.user, VopResultado, self.queryset, self.request
        )
